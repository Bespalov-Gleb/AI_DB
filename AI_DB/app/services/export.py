from __future__ import annotations
from pathlib import Path
from typing import Iterable, List, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.models.listings import Listing
from app.models.audit_log import AuditLog
from sqlalchemy.orm import Session
from decimal import Decimal
import json
from app.services.text_normalizer import normalize_contact


def _translate_columns_to_russian(df: pd.DataFrame) -> pd.DataFrame:
    """Переводит названия колонок DataFrame на русский язык"""
    column_mapping = {
        # Основные поля
        "id": "ID",
        "type": "Тип",
        "title": "Наименование",
        "description": "Описание",
        "characteristics": "Характеристики",
        "quantity": "Количество",
        "price": "Цена",
        "location": "Город",
        "contact": "Контакты",
        "photo_links": "Фотографии",
        "created_at": "Дата создания",
        "updated_at": "Дата обновления",
        
        # Поля для совпадений
        "score": "Оценка совпадения",
        "Demand": "Спрос",
        "Sale": "Предложение",
        
        # Детальные поля для совпадений
        "demand_id": "ID спроса",
        "demand_title": "Наименование спроса",
        "demand_location": "Город спроса",
        "demand_price": "Цена спроса",
        "demand_contact": "Контакты спроса",
        "sale_id": "ID предложения",
        "sale_title": "Наименование предложения",
        "sale_location": "Город предложения",
        "sale_price": "Цена предложения",
        "sale_contact": "Контакты предложения",
        
        # Поля аудита
        "actor": "Пользователь",
        "action": "Действие",
        "resource": "Ресурс",
        "result": "Результат",
        "payload": "Детали",
        
        # Статистика
        "count": "Количество"
    }
    
    # Переименовываем колонки
    df = df.rename(columns=column_mapping)
    return df


def export_listings_to_excel(
	listings: Iterable[Listing],
	filepath: str | Path,
	listing_id_to_photos: Dict[int, List[str]] | None = None,
) -> str:
	rows = []
	for l in listings:
		photo_links_list = []
		if listing_id_to_photos is not None:
			photo_links_list = listing_id_to_photos.get(l.id, [])
		if not photo_links_list:
			photo_links_list = l.photo_links or []
		rows.append({
			"id": l.id,
			"type": l.type,
			"title": l.title,
			"description": l.description,
			"characteristics": l.characteristics,
			"quantity": l.quantity,
			"price": float(l.price) if l.price is not None else None,
			"location": l.location,
			"contact": l.contact,
			"photo_links": ",".join(photo_links_list),
			"created_at": l.created_at,
			"updated_at": l.updated_at,
		})
	df = pd.DataFrame(rows)
	
	# Переводим колонки на русский
	df = _translate_columns_to_russian(df)
	
	filepath = str(filepath)
	df.to_excel(filepath, index=False)

	_auto_fit_excel(filepath)
	return filepath


def export_matches_to_excel(matches: List[dict], filepath: str | Path) -> str:
	df = pd.DataFrame(matches)
	
	# Переводим колонки на русский
	df = _translate_columns_to_russian(df)
	
	filepath = str(filepath)
	df.to_excel(filepath, index=False)
	_auto_fit_excel(filepath)
	return filepath


def export_stats_to_excel(listings: List[Listing], filepath: str | Path) -> str:
	# Лист 1: агрегаты по типу
	by_type = pd.DataFrame([{ "type": l.type or "", "count": 1 } for l in listings])
	by_type = by_type.groupby("type", as_index=False).sum()
	# Лист 2: агрегаты по городу
	by_city = pd.DataFrame([{ "location": (l.location or "").strip() or "(пусто)", "count": 1 } for l in listings])
	by_city = by_city.groupby("location", as_index=False).sum().sort_values("count", ascending=False)
	# Лист 3: сырой список (ограниченный набор полей)
	raw = pd.DataFrame([
		{ "id": l.id, "type": l.type, "title": l.title, "price": float(l.price) if l.price is not None else None, "location": l.location, "created_at": l.created_at }
		for l in listings
	])

	# Переводим колонки на русский для каждого листа
	by_type = _translate_columns_to_russian(by_type)
	by_city = _translate_columns_to_russian(by_city)
	raw = _translate_columns_to_russian(raw)

	filepath = str(filepath)
	with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
		by_type.to_excel(writer, index=False, sheet_name="По типу")
		by_city.to_excel(writer, index=False, sheet_name="По городу")
		raw.to_excel(writer, index=False, sheet_name="Список")

	# Автоподбор для каждого листа
	wb = load_workbook(filepath)
	for ws in wb.worksheets:
		for col_cells in ws.iter_cols(1, ws.max_column):
			max_len = 0
			for cell in col_cells:
				value = cell.value
				if value is None:
					continue
				try:
					val_str = str(value)
				except Exception:
					val_str = ""
				if len(val_str) > max_len:
					max_len = len(val_str)
			col_letter = get_column_letter(col_cells[0].column)
			ws.column_dimensions[col_letter].width = min(60, max_len + 2)
	wb.save(filepath)
	return filepath


def export_audit_to_excel(audit: List[AuditLog], filepath: str | Path) -> str:
	rows = []
	for a in audit:
		rows.append({
			"id": a.id,
			"created_at": a.created_at,
			"actor": a.actor,
			"action": a.action,
			"resource": a.resource,
			"result": a.result,
			"payload": a.payload,
		})
	df = pd.DataFrame(rows)
	
	# Переводим колонки на русский
	df = _translate_columns_to_russian(df)
	
	filepath = str(filepath)
	df.to_excel(filepath, index=False)
	_auto_fit_excel(filepath)
	return filepath


def _auto_fit_excel(filepath: str) -> None:
	wb = load_workbook(filepath)
	ws = wb.active
	for col_cells in ws.iter_cols(1, ws.max_column):
		max_len = 0
		for cell in col_cells:
			value = cell.value
			if value is None:
				continue
			try:
				val_str = str(value)
			except Exception:
				val_str = ""
			if len(val_str) > max_len:
				max_len = len(val_str)
		col_letter = get_column_letter(col_cells[0].column)
		ws.column_dimensions[col_letter].width = min(60, max_len + 2)
	wb.save(filepath)


def import_listings_from_excel(session: Session, filepath: str | Path) -> int:
	"""Импортирует записи из Excel того же формата, что и экспорт, и сливает с БД.
	Правила слияния:
	- Если указан id и запись существует — обновляем поля.
	- Если id пуст / не найден — создаём новую запись.
	Возвращает количество обработанных строк.
	"""
	path = str(filepath)
	df = pd.read_excel(path)
	count = 0
	for _, row in df.iterrows():
		try:
			rid = int(row.get("id")) if pd.notna(row.get("id")) else None
		except Exception:
			rid = None
		item: Listing | None = session.get(Listing, rid) if rid else None
		def _val(col):
			v = row.get(col)
			return None if (pd.isna(v) if hasattr(pd, "isna") else v is None) else v
		def _to_int(v):
			if v is None or v == "":
				return None
			try:
				return int(float(str(v).replace(" ", "")))
			except Exception:
				return None
		def _to_dec(v):
			if v is None or v == "":
				return None
			try:
				return Decimal(str(v).replace(" ", ""))
			except Exception:
				return None
		def _to_json(v):
			if v is None or v == "":
				return None
			try:
				return json.loads(v) if isinstance(v, str) else v
			except Exception:
				return None
		if item is None:
			quantity_val = _to_int(_val("quantity"))
			print(f"DEBUG: Row {count}: title={_val('title')}, quantity={_val('quantity')}->{quantity_val}, location={_val('location')}")
			print(f"DEBUG: All columns: {dict(row)}")
			item = Listing(
				title=str(_val("title") or "").strip() or "Без названия",
				description=_val("description"),
				characteristics=_to_json(_val("characteristics")),
				quantity=quantity_val,
				price=_to_dec(_val("price")),
				location=_val("location"),
				contact=normalize_contact(_val("contact")),
				photo_links=(str(_val("photo_links")) or "").split(",") if _val("photo_links") else None,
				type=str(_val("type") or "demand"),
			)
			session.add(item)
		else:
			item.title = str(_val("title") or item.title)
			item.description = _val("description")
			item.characteristics = _to_json(_val("characteristics"))
			item.quantity = _to_int(_val("quantity"))
			item.price = _to_dec(_val("price"))
			item.location = _val("location")
			item.contact = normalize_contact(_val("contact"))
			item.photo_links = (str(_val("photo_links")) or "").split(",") if _val("photo_links") else item.photo_links
			item.type = str(_val("type") or item.type)
		count += 1
		# Принудительно коммитим каждую запись отдельно, чтобы избежать проблем с bulk insert
		session.flush()
	session.commit()
	return count